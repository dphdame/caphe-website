#!/usr/bin/env python3
"""
Enrich _summary.json with per-specialty data from existing county JSON files
and/or by extracting specialty breakdowns from NPPES source data.

This script:
1. Reads existing county JSON files in data/access-explorer/
2. Extracts per-specialty {participationRate, registered, active} from each
3. Merges specialty data into _summary.json
4. Recomputes HRR-level specialty aggregates

For counties without individual JSON files, it attempts to stream-filter
the NPPES source file (reading CA rows only) and cross-reference with
spending data to compute specialty breakdowns.
"""

import csv
import json
import os
import sys
from pathlib import Path

SPECIALTY_KEYS = [
    "primary_care", "behavioral_health", "dental",
    "obgyn", "other_surgical", "pharmacy_dme"
]

TAXONOMY_MAP = {
    "primary_care": [
        "207Q00000X", "207QA0505X", "208D00000X", "207R00000X",
        "208000000X", "207RG0100X", "363L00000X", "363LA2200X",
        "363LF0000X", "363LP0200X",
    ],
    "behavioral_health": [
        "2084P0800X", "103T00000X", "103TA0400X", "103TC0700X",
        "104100000X", "1041C0700X", "106H00000X", "101YM0800X",
        "101YA0400X", "363LP0808X",
    ],
    "dental": [
        "1223G0001X", "1223P0221X", "1223S0112X", "1223X0400X",
        "1223E0200X", "1223P0106X", "124Q00000X",
    ],
    "obgyn": [
        "207V00000X", "207VB0002X", "207VM0101X", "207VX0201X",
        "176B00000X",
    ],
    "other_surgical": [
        "208600000X", "207X00000X", "207W00000X", "207Y00000X",
        "208800000X", "207T00000X", "208G00000X",
    ],
    "pharmacy_dme": [
        "183500000X", "3336C0003X", "332B00000X", "335E00000X",
    ],
}

# Invert for lookup
TAXONOMY_LOOKUP = {}
for specialty, codes in TAXONOMY_MAP.items():
    for code in codes:
        TAXONOMY_LOOKUP[code] = specialty

# HUD crosswalk California FIPS
CA_COUNTY_FIPS = {
    "06001": "Alameda", "06003": "Alpine", "06005": "Amador",
    "06007": "Butte", "06009": "Calaveras", "06011": "Colusa",
    "06013": "Contra Costa", "06015": "Del Norte", "06017": "El Dorado",
    "06019": "Fresno", "06021": "Glenn", "06023": "Humboldt",
    "06025": "Imperial", "06027": "Inyo", "06029": "Kern",
    "06031": "Kings", "06033": "Lake", "06035": "Lassen",
    "06037": "Los Angeles", "06039": "Madera", "06041": "Marin",
    "06043": "Mariposa", "06045": "Mendocino", "06047": "Merced",
    "06049": "Modoc", "06051": "Mono", "06053": "Monterey",
    "06055": "Napa", "06057": "Nevada", "06059": "Orange",
    "06061": "Placer", "06063": "Plumas", "06065": "Riverside",
    "06067": "Sacramento", "06069": "San Benito", "06071": "San Bernardino",
    "06073": "San Diego", "06075": "San Francisco", "06077": "San Joaquin",
    "06079": "San Luis Obispo", "06081": "San Mateo",
    "06083": "Santa Barbara", "06085": "Santa Clara", "06087": "Santa Cruz",
    "06089": "Shasta", "06091": "Sierra", "06093": "Siskiyou",
    "06095": "Solano", "06097": "Sonoma", "06099": "Stanislaus",
    "06101": "Sutter", "06103": "Tehama", "06105": "Trinity",
    "06107": "Tulare", "06109": "Tuolumne", "06111": "Ventura",
    "06113": "Yolo", "06115": "Yuba",
}


def enrich_from_county_files(data_dir, summary):
    """Read existing county JSON files and extract per-specialty data."""
    enriched = 0
    for county_name in summary["counties"]:
        file_name = county_name.lower().replace(" ", "_") + ".json"
        file_path = os.path.join(data_dir, file_name)
        if not os.path.exists(file_path):
            continue

        with open(file_path) as f:
            county_data = json.load(f)

        specialties = {}
        for key in SPECIALTY_KEYS:
            spec = county_data.get("specialties", {}).get(key)
            if spec:
                specialties[key] = {
                    "participationRate": spec["participationRate"],
                    "registered": spec["registered"],
                    "active": spec["active"],
                }

        if specialties:
            summary["counties"][county_name]["specialties"] = specialties
            enriched += 1

    return enriched


def load_zip_county_crosswalk(crosswalk_path):
    """Load ZIP-to-county crosswalk. Supports .xlsx via openpyxl or .csv."""
    zip_to_county = {}

    if crosswalk_path.endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(crosswalk_path, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            zip_idx = headers.index("ZIP")
            county_idx = headers.index("COUNTY")
            ratio_idx = headers.index("RES_RATIO")

            # Keep highest RES_RATIO per ZIP
            best = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                z = str(row[zip_idx]).zfill(5)
                c = str(row[county_idx]).zfill(5)
                r = float(row[ratio_idx]) if row[ratio_idx] else 0
                if z not in best or r > best[z][1]:
                    best[z] = (c, r)
            wb.close()

            for z, (c, _) in best.items():
                county_name = CA_COUNTY_FIPS.get(c)
                if county_name:
                    zip_to_county[z] = county_name
        except ImportError:
            print("  openpyxl not available, trying polars...")
            import polars as pl
            xwalk = pl.read_excel(crosswalk_path)
            xwalk = xwalk.sort("RES_RATIO", descending=True).group_by("ZIP").first()
            for row in xwalk.iter_rows(named=True):
                z = str(row["ZIP"]).zfill(5)
                c = str(row["COUNTY"]).zfill(5)
                county_name = CA_COUNTY_FIPS.get(c)
                if county_name:
                    zip_to_county[z] = county_name
    else:
        with open(crosswalk_path) as f:
            reader = csv.DictReader(f)
            best = {}
            for row in reader:
                z = row["ZIP"].zfill(5)
                c = row["COUNTY"].zfill(5)
                r = float(row.get("RES_RATIO", 0))
                if z not in best or r > best[z][1]:
                    best[z] = (c, r)
            for z, (c, _) in best.items():
                county_name = CA_COUNTY_FIPS.get(c)
                if county_name:
                    zip_to_county[z] = county_name

    print(f"  ZIP-county mappings loaded: {len(zip_to_county):,}")
    return zip_to_county


def stream_nppes_ca(nppes_path, zip_to_county):
    """
    Stream NPPES CSV, filter CA providers, classify by specialty.
    Returns dict: county -> specialty -> list of NPIs
    """
    county_spec_npis = {}
    processed = 0
    matched = 0

    print(f"  Streaming NPPES from {nppes_path}...")

    with open(nppes_path, 'r', errors='replace') as f:
        reader = csv.DictReader(f)
        for row in reader:
            processed += 1
            if processed % 1_000_000 == 0:
                print(f"    Processed {processed:,} rows, matched {matched:,} CA providers...")

            # Filter: CA only, active NPI
            state = row.get("Provider Business Practice Location Address State Name", "")
            if state != "CA":
                continue
            if row.get("NPI Deactivation Date"):
                continue

            # Get ZIP -> county
            zip5 = (row.get("Provider Business Practice Location Address Postal Code") or "")[:5]
            county = zip_to_county.get(zip5)
            if not county:
                continue

            # Classify specialty
            taxonomy = row.get("Healthcare Provider Taxonomy Code_1", "")
            specialty = TAXONOMY_LOOKUP.get(taxonomy)
            if not specialty:
                continue

            npi = row.get("NPI", "")
            if not npi:
                continue

            matched += 1
            if county not in county_spec_npis:
                county_spec_npis[county] = {}
            if specialty not in county_spec_npis[county]:
                county_spec_npis[county][specialty] = []
            county_spec_npis[county][specialty].append(npi)

    print(f"  Total processed: {processed:,}, CA matched: {matched:,}")
    return county_spec_npis


def load_active_npis(spending_path):
    """Load active NPIs from spending data."""
    active = set()
    with open(spending_path, 'r', errors='replace') as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("Rndrng_Prvdr_State_Abrvtn") == "CA":
                npi = row.get("Rndrng_NPI", "")
                if npi:
                    active.add(npi)
    print(f"  Active CA Medicaid NPIs: {len(active):,}")
    return active


def compute_specialty_breakdowns(county_spec_npis, active_npis):
    """Compute per-county per-specialty participation rates."""
    result = {}
    for county, specs in county_spec_npis.items():
        result[county] = {}
        for spec_key, npis in specs.items():
            registered = len(npis)
            active = sum(1 for npi in npis if npi in active_npis)
            rate = round(active / registered * 100, 1) if registered > 0 else 0
            result[county][spec_key] = {
                "participationRate": rate,
                "registered": registered,
                "active": active,
            }
    return result


def recompute_hrr_specialties(summary):
    """Recompute HRR-level specialty aggregates from county data."""
    if "hrrs" not in summary:
        return

    for hrr_name, hrr_data in summary["hrrs"].items():
        hrr_specialties = {}
        for spec_key in SPECIALTY_KEYS:
            spec_reg = 0
            spec_active = 0
            for county_name in hrr_data.get("counties", []):
                county = summary["counties"].get(county_name, {})
                spec_data = county.get("specialties", {}).get(spec_key, {})
                spec_reg += spec_data.get("registered", 0)
                spec_active += spec_data.get("active", 0)
            if spec_reg > 0:
                hrr_specialties[spec_key] = {
                    "participationRate": round(spec_active / spec_reg * 100, 1),
                    "registered": spec_reg,
                    "active": spec_active,
                }
        hrr_data["specialties"] = hrr_specialties

    print(f"  Recomputed specialties for {len(summary['hrrs'])} HRRs")


def main():
    data_dir = os.path.join(os.path.dirname(__file__), "..", "data", "access-explorer")
    summary_path = os.path.join(data_dir, "_summary.json")

    with open(summary_path) as f:
        summary = json.load(f)

    total_counties = len(summary["counties"])
    print(f"Summary has {total_counties} counties")

    # Phase 1: Enrich from existing county JSON files
    enriched = enrich_from_county_files(data_dir, summary)
    print(f"Enriched {enriched} counties from existing JSON files")

    # Check how many still need enrichment
    missing = [
        name for name in summary["counties"]
        if "specialties" not in summary["counties"][name]
    ]
    print(f"Counties still missing specialty data: {len(missing)}")

    if missing:
        # Phase 2: Stream NPPES to fill gaps
        nppes_path = os.environ.get("NPPES_PATH")
        spending_path = os.environ.get("SPENDING_PATH")
        crosswalk_path = os.environ.get("CROSSWALK_PATH")

        if nppes_path and spending_path and crosswalk_path:
            print("\nPhase 2: Loading source data for remaining counties...")
            zip_to_county = load_zip_county_crosswalk(crosswalk_path)
            county_spec_npis = stream_nppes_ca(nppes_path, zip_to_county)
            active_npis = load_active_npis(spending_path)
            breakdowns = compute_specialty_breakdowns(county_spec_npis, active_npis)

            for county_name in missing:
                if county_name in breakdowns:
                    summary["counties"][county_name]["specialties"] = breakdowns[county_name]

            still_missing = [
                name for name in summary["counties"]
                if "specialties" not in summary["counties"][name]
            ]
            print(f"After NPPES processing: {len(still_missing)} counties still missing")
        else:
            print("\nSet NPPES_PATH, SPENDING_PATH, CROSSWALK_PATH to fill gaps from source data")
            print("Proceeding with partial enrichment...")

    # Phase 3: Recompute HRR specialties
    recompute_hrr_specialties(summary)

    # Write enriched summary
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)

    # Report
    enriched_count = sum(
        1 for c in summary["counties"].values() if "specialties" in c
    )
    print(f"\nDone! {enriched_count}/{total_counties} counties have specialty data")
    print(f"Written to {summary_path}")


if __name__ == "__main__":
    main()
